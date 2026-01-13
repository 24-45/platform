#!/usr/bin/env python3
"""
تصدير خطة الحملة الإعلامية إلى ملف Excel
Export Campaign Media Plan to Excel
"""

import json
import os
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("جاري تثبيت مكتبة openpyxl...")
    os.system('pip install openpyxl')
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# مسارات الملفات
BASE_PATH = Path(__file__).parent
DATA_PATH = BASE_PATH / 'data' / 'campaigns' / 'campaigns.json'
OUTPUT_PATH = BASE_PATH / 'ALIC_Campaign_Plan.xlsx'

# ألوان المراحل
PHASE_COLORS = {
    'teasing': 'FFF59E0B',      # أصفر - مرحلة التشويق
    'launch': 'FF3B82F6',       # أزرق - مرحلة الإطلاق
    'momentum': 'FF10B981',     # أخضر - مرحلة الترسيخ
}

PHASE_NAMES = {
    'teasing': 'مرحلة التشويق (Teasing)',
    'launch': 'مرحلة الإطلاق (Launch)',
    'momentum': 'مرحلة الترسيخ (Momentum)',
}

def load_campaign_data():
    """تحميل بيانات الحملة"""
    with open(DATA_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # البحث عن حملة ALIC
    for campaign in data.get('campaigns', []):
        if campaign.get('id') == 'alic-almuwaqqar':
            return campaign
    return None

def create_excel_report(campaign):
    """إنشاء تقرير Excel"""
    wb = openpyxl.Workbook()
    
    # إزالة الشيت الافتراضي
    wb.remove(wb.active)
    
    # ========== شيت 1: نظرة عامة ==========
    ws_overview = wb.create_sheet("نظرة عامة", 0)
    create_overview_sheet(ws_overview, campaign)
    
    # ========== شيت 2: المنتجات حسب المراحل ==========
    ws_products = wb.create_sheet("المنتجات والمراحل", 1)
    create_products_sheet(ws_products, campaign)
    
    # ========== شيت 3: الأدوات الإعلامية ==========
    ws_tools = wb.create_sheet("الأدوات الإعلامية", 2)
    create_tools_sheet(ws_tools, campaign)
    
    # ========== شيت 4: الميزانية ==========
    ws_budget = wb.create_sheet("الميزانية", 3)
    create_budget_sheet(ws_budget, campaign)
    
    # حفظ الملف
    wb.save(OUTPUT_PATH)
    print(f"\n✅ تم إنشاء ملف Excel بنجاح:")
    print(f"   📁 {OUTPUT_PATH}")
    return OUTPUT_PATH

def apply_header_style(cell, color='FF1E3A5F'):
    """تطبيق تنسيق رأس الجدول"""
    cell.font = Font(bold=True, color='FFFFFF', size=12)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def apply_cell_style(cell, wrap=True):
    """تطبيق تنسيق الخلية العادية"""
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=wrap)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def create_overview_sheet(ws, campaign):
    """إنشاء شيت النظرة العامة"""
    ws.sheet_view.rightToLeft = True
    
    # العنوان الرئيسي
    ws.merge_cells('A1:D1')
    ws['A1'] = f"خطة حملة {campaign['basic_info']['name']}"
    ws['A1'].font = Font(bold=True, size=18, color='DC1F27')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # معلومات الحملة
    info_data = [
        ('اسم الحملة', campaign['basic_info']['name']),
        ('اسم الحملة (إنجليزي)', campaign['basic_info'].get('name_en', '')),
        ('العميل', campaign['client_info']['company_name']),
        ('المشروع', campaign['project_info']['project_name']),
        ('تاريخ البداية', campaign['basic_info'].get('start_date', '')),
        ('تاريخ النهاية', campaign['basic_info'].get('end_date', '')),
        ('مدة الحملة', f"{campaign['basic_info'].get('duration_days', 0)} يوم"),
        ('إجمالي المنتجات', campaign['basic_info'].get('total_products', 0)),
        ('الميزانية الإجمالية', f"{campaign['basic_info'].get('budget', 0):,} {campaign['basic_info'].get('currency', 'USD')}"),
    ]
    
    row = 3
    for label, value in info_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    # تعديل عرض الأعمدة
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

def create_products_sheet(ws, campaign):
    """إنشاء شيت المنتجات حسب المراحل"""
    ws.sheet_view.rightToLeft = True
    
    # العناوين
    headers = ['#', 'المرحلة', 'اسم المنتج', 'الفئة', 'النوع', 'التاريخ', 'الوصف', 'الحالة', 'التكلفة']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
    
    # المنتجات
    products = campaign.get('communication_products', [])
    row = 2
    
    for i, product in enumerate(products, 1):
        phase_id = product.get('phase_id', '')
        phase_name = PHASE_NAMES.get(phase_id, phase_id)
        phase_color = PHASE_COLORS.get(phase_id, 'FFFFFFFF')
        
        data = [
            i,
            phase_name,
            product.get('name', ''),
            product.get('category', ''),
            product.get('type', ''),
            product.get('date', ''),
            product.get('description', ''),
            'قيد الانتظار' if product.get('status') == 'pending' else product.get('status', ''),
            product.get('cost', 0)
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
            
            # تلوين خلية المرحلة
            if col == 2:
                cell.fill = PatternFill(start_color=phase_color, end_color=phase_color, fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
        
        row += 1
    
    # ملخص المراحل
    row += 2
    ws.cell(row=row, column=1, value="ملخص المراحل")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 1
    
    phase_summary = {}
    for product in products:
        phase = product.get('phase_id', 'other')
        if phase not in phase_summary:
            phase_summary[phase] = {'count': 0, 'cost': 0}
        phase_summary[phase]['count'] += 1
        phase_summary[phase]['cost'] += product.get('cost', 0)
    
    for phase_id, summary in phase_summary.items():
        phase_name = PHASE_NAMES.get(phase_id, phase_id)
        ws.cell(row=row, column=1, value=phase_name)
        ws.cell(row=row, column=2, value=f"{summary['count']} منتج")
        ws.cell(row=row, column=3, value=f"{summary['cost']:,} USD")
        row += 1
    
    # تعديل عرض الأعمدة
    column_widths = [5, 30, 50, 15, 20, 20, 40, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def create_tools_sheet(ws, campaign):
    """إنشاء شيت الأدوات الإعلامية"""
    ws.sheet_view.rightToLeft = True
    
    # العناوين
    headers = ['#', 'المرحلة', 'اسم الأداة', 'النوع', 'مدفوع/مجاني', 'التكلفة', 'ملاحظات']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell, 'FFEC4899')  # لون وردي
    
    # الأدوات
    tools = campaign.get('media_tools', [])
    row = 2
    
    for i, tool in enumerate(tools, 1):
        phase_id = tool.get('phase_id', '')
        phase_name = PHASE_NAMES.get(phase_id, phase_id)
        
        data = [
            i,
            phase_name,
            tool.get('name', ''),
            tool.get('type', ''),
            'مدفوع 💰' if tool.get('is_paid') else 'مجاني ✓',
            f"{tool.get('cost', 0):,} USD" if tool.get('cost', 0) > 0 else '-',
            tool.get('notes', '')
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
            
            # تلوين حسب نوع الدفع
            if col == 5:
                if tool.get('is_paid'):
                    cell.fill = PatternFill(start_color='FFFEF3C7', end_color='FFFEF3C7', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFD1FAE5', end_color='FFD1FAE5', fill_type='solid')
        
        row += 1
    
    # ملخص التكاليف
    row += 2
    ws.cell(row=row, column=1, value="ملخص تكاليف الأدوات")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 1
    
    total_paid = sum(t.get('cost', 0) for t in tools if t.get('is_paid'))
    total_free = len([t for t in tools if not t.get('is_paid')])
    total_paid_count = len([t for t in tools if t.get('is_paid')])
    
    ws.cell(row=row, column=1, value="إجمالي الأدوات المدفوعة")
    ws.cell(row=row, column=2, value=f"{total_paid_count} أداة")
    ws.cell(row=row, column=3, value=f"{total_paid:,} USD")
    row += 1
    
    ws.cell(row=row, column=1, value="إجمالي الأدوات المجانية")
    ws.cell(row=row, column=2, value=f"{total_free} أداة")
    row += 1
    
    ws.cell(row=row, column=1, value="الإجمالي الكلي")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"{len(tools)} أداة")
    ws.cell(row=row, column=3, value=f"{total_paid:,} USD")
    ws.cell(row=row, column=3).font = Font(bold=True, color='DC1F27')
    
    # تعديل عرض الأعمدة
    column_widths = [5, 30, 35, 20, 15, 15, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def create_budget_sheet(ws, campaign):
    """إنشاء شيت الميزانية"""
    ws.sheet_view.rightToLeft = True
    
    budget = campaign.get('budget', {})
    
    # العنوان
    ws.merge_cells('A1:D1')
    ws['A1'] = "توزيع الميزانية"
    ws['A1'].font = Font(bold=True, size=16, color='1E3A5F')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # الميزانية الإجمالية
    ws['A3'] = "الميزانية الإجمالية:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = f"{budget.get('total', 0):,} {budget.get('currency', 'USD')}"
    ws['B3'].font = Font(bold=True, size=14, color='DC1F27')
    
    # العناوين
    headers = ['الفئة', 'المبلغ', 'النسبة']
    row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    
    # التفاصيل
    row = 6
    breakdown = budget.get('breakdown', [])
    for item in breakdown:
        ws.cell(row=row, column=1, value=item.get('category', ''))
        apply_cell_style(ws.cell(row=row, column=1))
        
        ws.cell(row=row, column=2, value=f"{item.get('amount', 0):,} USD")
        apply_cell_style(ws.cell(row=row, column=2))
        
        ws.cell(row=row, column=3, value=f"{item.get('percentage', 0)}%")
        apply_cell_style(ws.cell(row=row, column=3))
        
        row += 1
    
    # تعديل عرض الأعمدة
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

def main():
    """الدالة الرئيسية"""
    print("=" * 50)
    print("   📊 تصدير خطة الحملة إلى Excel")
    print("=" * 50)
    
    # تحميل البيانات
    print("\n⏳ جاري تحميل بيانات الحملة...")
    campaign = load_campaign_data()
    
    if not campaign:
        print("❌ لم يتم العثور على حملة ALIC")
        return
    
    print(f"✓ تم تحميل حملة: {campaign['basic_info']['name']}")
    print(f"  - عدد المنتجات: {len(campaign.get('communication_products', []))}")
    print(f"  - عدد الأدوات: {len(campaign.get('media_tools', []))}")
    
    # إنشاء التقرير
    print("\n⏳ جاري إنشاء ملف Excel...")
    output_path = create_excel_report(campaign)
    
    print("\n" + "=" * 50)
    print("   ✅ تم بنجاح!")
    print("=" * 50)

if __name__ == '__main__':
    main()
