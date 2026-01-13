#!/usr/bin/env python3
"""
تصدير خطة الحملة الإعلامية إلى ملف Excel - نسخة تفصيلية
Export Campaign Media Plan to Excel - Detailed Version
"""

import json
import os
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("جاري تثبيت مكتبة openpyxl...")
    os.system('pip install openpyxl')
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# مسارات الملفات
BASE_PATH = Path(__file__).parent
DATA_PATH = BASE_PATH / 'data' / 'campaigns' / 'campaigns.json'
OUTPUT_PATH = BASE_PATH / 'ALIC_Campaign_Plan_Detailed.xlsx'

# ألوان المراحل
PHASE_COLORS = {
    'teasing': 'FFF59E0B',      # أصفر - مرحلة التشويق
    'launch': 'FF3B82F6',       # أزرق - مرحلة الإطلاق
    'momentum': 'FF10B981',     # أخضر - مرحلة الترسيخ
}

PHASE_NAMES = {
    'teasing': 'مرحلة التشويق (Teasing)',
    'launch': 'مرحلة الإطلاق (Launch)',
    'momentum': 'مرحلة الترسيخ (Momentum)',
}

# ======== قوائم الأسماء التفصيلية ========

# الصحف الأردنية المحلية
NEWSPAPERS_LOCAL = [
    "صحيفة الرأي",
    "صحيفة الغد",
    "صحيفة الدستور",
    "صحيفة الأنباط",
    "صحيفة السبيل"
]

# المواقع الإخبارية الإلكترونية
NEWS_WEBSITES = [
    "موقع عمون",
    "رؤيا نيوز",
    "خبرني",
    "جفرا نيوز",
    "الوكيل الإخباري",
    "سرايا نيوز",
    "بترا - وكالة الأنباء الأردنية",
    "المدينة نيوز",
    "الأردن 24",
    "نيوز بلس"
]

# القنوات التلفزيونية الأردنية
TV_CHANNELS = [
    "قناة رؤيا",
    "قناة المملكة",
    "التلفزيون الأردني",
    "قناة الأردن اليوم",
    "قناة هلا"
]

# الإذاعات الأردنية
RADIO_STATIONS = [
    "إذاعة حسنى FM 100.3",
    "إذاعة روتانا FM",
    "إذاعة هلا FM 104.9",
    "إذاعة مزاج FM",
    "إذاعة فرح الناس FM"
]

# المؤثرين - مرحلة التشويق
INFLUENCERS_TEASING = [
    "أحمد حسن الزعبي (@ahmedzubi) - عقارات واستثمار",
    "د. نادر عزام (@naderazam) - اقتصاد وأعمال",
    "محمد البشير (@m_albasheer) - صناعة وتجارة"
]

# المؤثرين - مرحلة الإطلاق (25 شخص)
INFLUENCERS_LAUNCH = [
    "عمر العبداللات - مذيع ومقدم برامج",
    "رنا الحسين - إعلامية اقتصادية",
    "د. موسى شتيوي - خبير اقتصادي",
    "فادي غندور - رائد أعمال",
    "محمد حدادين - صحفي اقتصادي",
    "سامر طويل - مؤثر أعمال",
    "نسرين زريقات - إعلامية",
    "خالد خريس - صحفي",
    "د. إبراهيم سيف - محلل اقتصادي",
    "ياسر أبو هلالة - إعلامي"
]

# المؤثرين - مرحلة الترسيخ
INFLUENCERS_MOMENTUM = [
    "المخبر الاقتصادي (YouTube) - 5.5M مشترك",
    "طارق الخضراوي - مؤثر استثمار",
    "د. يوسف منصور - خبير عقارات",
    "رائد العبادي - صحفي اقتصادي",
    "سهير جرار - إعلامية"
]

# وسائل الإعلام الدولية
INTERNATIONAL_MEDIA = [
    "Reuters",
    "Bloomberg Middle East",
    "Arab News",
    "Gulf Business"
]

# البودكاست
PODCASTS = [
    "بودكاست سوالف بزنس",
    "بودكاست فنجان (مع ثمود بن محفوظ)",
    "بودكاست أبجد هوز"
]


def load_campaign_data():
    """تحميل بيانات الحملة"""
    with open(DATA_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    for campaign in data.get('campaigns', []):
        if campaign.get('id') == 'alic-almuwaqqar':
            return campaign
    return None

def create_excel_report(campaign):
    """إنشاء تقرير Excel"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # شيت 1: نظرة عامة
    ws_overview = wb.create_sheet("نظرة عامة", 0)
    create_overview_sheet(ws_overview, campaign)
    
    # شيت 2: خطة المنتجات التفصيلية
    ws_products = wb.create_sheet("خطة المنتجات التفصيلية", 1)
    create_detailed_products_sheet(ws_products, campaign)
    
    # شيت 3: الأدوات الإعلامية بالأسماء
    ws_tools = wb.create_sheet("الأدوات الإعلامية التفصيلية", 2)
    create_detailed_tools_sheet(ws_tools, campaign)
    
    # شيت 4: الوسائل المدفوعة
    ws_paid = wb.create_sheet("الوسائل المدفوعة", 3)
    create_paid_media_sheet(ws_paid, campaign)
    
    # شيت 5: الإعلام الموازي (Earned)
    ws_earned = wb.create_sheet("الإعلام الموازي", 4)
    create_earned_media_sheet(ws_earned, campaign)
    
    # شيت 6: الميزانية
    ws_budget = wb.create_sheet("الميزانية", 5)
    create_budget_sheet(ws_budget, campaign)
    
    wb.save(OUTPUT_PATH)
    print(f"\n✅ تم إنشاء ملف Excel بنجاح:")
    print(f"   📁 {OUTPUT_PATH}")
    return OUTPUT_PATH

def apply_header_style(cell, color='FF1E3A5F'):
    """تطبيق تنسيق رأس الجدول"""
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def apply_cell_style(cell, wrap=True):
    """تطبيق تنسيق الخلية العادية"""
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=wrap)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def create_overview_sheet(ws, campaign):
    """إنشاء شيت النظرة العامة"""
    ws.sheet_view.rightToLeft = True
    
    ws.merge_cells('A1:D1')
    ws['A1'] = f"📊 خطة حملة {campaign['basic_info']['name']} - التفصيلية"
    ws['A1'].font = Font(bold=True, size=18, color='DC1F27')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    info_data = [
        ('اسم الحملة', campaign['basic_info']['name']),
        ('العميل', campaign['client_info']['company_name']),
        ('المشروع', campaign['project_info']['project_name']),
        ('الموقع', campaign['project_info'].get('project_location', 'جنوب عمّان')),
        ('تاريخ البداية', campaign['basic_info'].get('start_date', '')),
        ('تاريخ النهاية', campaign['basic_info'].get('end_date', '')),
        ('مدة الحملة', f"{campaign['basic_info'].get('duration_days', 0)} يوم"),
        ('إجمالي المنتجات', f"{campaign['basic_info'].get('total_products', 0)} منتج"),
        ('الميزانية الإجمالية', f"{campaign['basic_info'].get('budget', 0):,} {campaign['basic_info'].get('currency', 'USD')}"),
    ]
    
    row = 3
    for label, value in info_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    # ملخص المراحل
    row += 2
    ws.cell(row=row, column=1, value="📅 مراحل الحملة")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color='3B82F6')
    row += 1
    
    phases_info = [
        ('مرحلة التشويق (Teasing)', '5-12 يناير 2026', '8 أيام', '5 منتجات'),
        ('مرحلة الإطلاق (Launch)', '13-22 يناير 2026', '10 أيام', '10 منتجات'),
        ('مرحلة الترسيخ (Momentum)', '23-31 يناير 2026', '9 أيام', '5 منتجات'),
    ]
    
    for phase in phases_info:
        ws.cell(row=row, column=1, value=phase[0])
        ws.cell(row=row, column=2, value=phase[1])
        ws.cell(row=row, column=3, value=phase[2])
        ws.cell(row=row, column=4, value=phase[3])
        row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

def create_detailed_products_sheet(ws, campaign):
    """إنشاء شيت المنتجات التفصيلي مع الأدوات المرتبطة"""
    ws.sheet_view.rightToLeft = True
    
    # الحصول على العملة من العرض المالي
    quotation = campaign.get('quotation', {})
    currency = quotation.get('currency', 'USD')
    
    headers = ['#', 'المرحلة', 'اسم المنتج', 'النوع', 'التاريخ', 'الأدوات الاتصالية المستخدمة', 'الكمية', f'سعر الوحدة ({currency})', f'الإجمالي ({currency})']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
    
    products_list = quotation.get('products_list', campaign.get('communication_products', []))
    
    row = 2
    for i, product in enumerate(products_list, 1):
        phase_id = product.get('phase', product.get('phase_id', ''))
        phase_name = PHASE_NAMES.get(phase_id, phase_id)
        phase_color = PHASE_COLORS.get(phase_id, 'FFFFFFFF')
        
        # تحديد الأدوات الاتصالية حسب المنتج
        tools = get_tools_for_product(product)
        
        quantity = product.get('quantity', 1)
        unit_price = product.get('unit_price', product.get('cost', 0))
        total = quantity * unit_price
        
        data = [
            i,
            phase_name,
            product.get('name', ''),
            product.get('type', ''),
            product.get('date', ''),
            tools,
            quantity,
            f"{unit_price:,}",
            f"{total:,}"
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
            
            if col == 2:
                cell.fill = PatternFill(start_color=phase_color, end_color=phase_color, fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
        
        row += 1
    
    # ملخص المنتجات مع الإجمالي
    row += 2
    ws.cell(row=row, column=1, value="📊 ملخص المنتجات")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 1
    
    # حساب الإجمالي الكلي
    grand_total = sum((p.get('quantity', 1) * p.get('unit_price', 0)) for p in products_list)
    ws.cell(row=row, column=1, value="إجمالي تكاليف المنتجات")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=9, value=f"{grand_total:,} {currency}")
    ws.cell(row=row, column=9).font = Font(bold=True, color='DC1F27', size=12)
    
    column_widths = [5, 28, 45, 18, 20, 60, 10, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def get_tools_for_product(product):
    """تحديد الأدوات الاتصالية لكل منتج"""
    name = product.get('name', '').lower()
    product_type = product.get('type', '').lower()
    
    tools_mapping = {
        'linkedin': 'LinkedIn Ads Manager, Canva Pro, Hootsuite',
        'بيان صحفي': f"PR Newswire, {', '.join(NEWSPAPERS_LOCAL[:3])}, {', '.join(NEWS_WEBSITES[:3])}",
        'sms': 'Twilio SMS, Mailchimp Email, HubSpot CRM',
        'email': 'Mailchimp Email, HubSpot CRM, SendGrid',
        'فيديو تشويقي': 'Adobe Premiere Pro, After Effects, YouTube Ads',
        'دعوات vip': 'Adobe InDesign, مطابع محلية, WhatsApp Business',
        'حفل الإطلاق': f'تنظيم فعاليات, {", ".join(TV_CHANNELS[:2])}, {", ".join(RADIO_STATIONS[:2])}',
        'مؤتمر صحفي': f'{", ".join(NEWSPAPERS_LOCAL)}, {", ".join(TV_CHANNELS[:3])}',
        'تغطية تلفزيونية': f'{", ".join(TV_CHANNELS)}',
        'إعلانات مدفوعة': 'Google Ads, Meta Ads Manager, LinkedIn Campaign Manager',
        'حملة linkedin': 'LinkedIn Ads Manager, Canva Pro, Sprout Social',
        'مؤثرين': f'{", ".join(INFLUENCERS_LAUNCH[:5])}',
        'بودكاست': f'{", ".join(PODCASTS)}',
        'مقالات': f'كتّاب رأي في: {", ".join(NEWSPAPERS_LOCAL)}',
        'جولة إعلامية': f'{", ".join(NEWSPAPERS_LOCAL[:3])}, {", ".join(TV_CHANNELS[:2])}, مصورين محترفين',
        '50+ وسيلة': f'PR Newswire, {", ".join(INTERNATIONAL_MEDIA)}',
        'إذاعي': f'{", ".join(RADIO_STATIONS)}',
        'جولة إذاعية': f'{", ".join(RADIO_STATIONS)}',
        'المخبر الاقتصادي': 'قناة المخبر الاقتصادي (YouTube 5.5M مشترك)',
        'مال وأعمال': f'{TV_CHANNELS[0]} - برنامج مال وأعمال',
        'شهادات': f'{", ".join(INFLUENCERS_MOMENTUM[:3])}, تصوير احترافي',
    }
    
    for key, tools in tools_mapping.items():
        if key in name:
            return tools
    
    return '-'

def create_detailed_tools_sheet(ws, campaign):
    """إنشاء شيت الأدوات الإعلامية التفصيلي"""
    ws.sheet_view.rightToLeft = True
    
    # الحصول على العملة والأدوات من العرض المالي
    quotation = campaign.get('quotation', {})
    currency = quotation.get('currency', 'USD')
    
    headers = ['#', 'المرحلة', 'نوع الأداة', 'الأسماء التفصيلية', 'مدفوع/مجاني', f'التكلفة ({currency})', 'ملاحظات']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell, 'FFEC4899')
    
    # قائمة الأدوات التفصيلية - الأسعار من العرض المالي
    media_tools = quotation.get('media_tools_list', campaign.get('media_tools', []))
    paid_media = quotation.get('paid_media_list', [])
    earned_media = quotation.get('earned_media_list', [])
    
    detailed_tools = [
        # مرحلة التشويق
        {
            'phase': 'teasing',
            'type': 'صحف محلية',
            'names': '\n'.join(NEWSPAPERS_LOCAL),
            'paid': False,
            'cost': 0,
            'notes': 'نشر بيانات صحفية - إعلام موازي'
        },
        {
            'phase': 'teasing',
            'type': 'مؤثرين (تشويق)',
            'names': '\n'.join(INFLUENCERS_TEASING),
            'paid': True,
            'cost': 500,  # من media_tools_list m2
            'notes': '3 مؤثرين في مجال الاستثمار والعقارات'
        },
        {
            'phase': 'teasing',
            'type': 'تواصل مباشر',
            'names': 'حملة SMS (5000 رسالة)\nحملة Email عبر Mailchimp\nقاعدة بيانات مستثمرين صناعيين',
            'paid': True,
            'cost': 750,  # من paid_media_list paid1
            'notes': 'استهداف مباشر للمستثمرين المحتملين'
        },
        {
            'phase': 'teasing',
            'type': 'إعلانات LinkedIn',
            'names': 'LinkedIn Campaign Manager\nاستهداف: صناعيين، لوجستيين، مستثمرين عقاريين',
            'paid': True,
            'cost': 1500,  # من paid_media_list paid2
            'notes': 'مرحلة التهيئة - بناء الوعي'
        },
        # مرحلة الإطلاق
        {
            'phase': 'launch',
            'type': 'صحافة دولية',
            'names': f'PR Newswire (50+ صحيفة):\n{chr(10).join(INTERNATIONAL_MEDIA)}\n+ صحف عربية ودولية متعددة',
            'paid': True,
            'cost': 1800,  # من paid_media_list paid6
            'notes': 'توزيع دولي للبيان الصحفي'
        },
        {
            'phase': 'launch',
            'type': 'قنوات تلفزيونية',
            'names': '\n'.join(TV_CHANNELS),
            'paid': True,
            'cost': 10500,  # من earned_media_list earned3 (3 × 3500)
            'notes': '3 قنوات - تغطيات إخبارية ولقاءات'
        },
        {
            'phase': 'launch',
            'type': 'إذاعات',
            'names': '\n'.join(RADIO_STATIONS),
            'paid': True,
            'cost': 2500,  # من earned_media_list earned4 (5 × 500)
            'notes': '5 مقابلات إذاعية'
        },
        {
            'phase': 'launch',
            'type': 'صحفيين ومؤثرين',
            'names': '\n'.join(INFLUENCERS_LAUNCH),
            'paid': True,
            'cost': 2500,  # من media_tools_list m7
            'notes': '25 صحفي ومؤثر للجولة الإعلامية'
        },
        {
            'phase': 'launch',
            'type': 'مؤثرين Micro (10K-50K)',
            'names': '15 مؤثر في مجالات: عقارات، استثمار، أعمال',
            'paid': True,
            'cost': 3750,  # من earned_media_list earned5 (15 × 250)
            'notes': 'مؤثرين صغار - انتشار واسع'
        },
        {
            'phase': 'launch',
            'type': 'مؤثرين Mid (50K-200K)',
            'names': '5 مؤثرين متوسطين في مجال الأعمال',
            'paid': True,
            'cost': 4000,  # من earned_media_list earned6 (5 × 800)
            'notes': 'مؤثرين متوسطين - تأثير قوي'
        },
        {
            'phase': 'launch',
            'type': 'إعلانات رقمية',
            'names': 'LinkedIn Ads: 3,000\nGoogle Display: 2,500\nMeta Ads: 2,000\nYouTube Pre-roll: 1,500',
            'paid': True,
            'cost': 9000,  # من paid3+paid4+paid5+paid8
            'notes': 'حملة إعلانية رقمية متكاملة'
        },
        {
            'phase': 'launch',
            'type': 'إعلام دولي',
            'names': '\n'.join(INTERNATIONAL_MEDIA),
            'paid': True,
            'cost': 2800,  # من paid_media_list paid7
            'notes': '4 وسائل إعلام دولية'
        },
        # مرحلة الترسيخ
        {
            'phase': 'momentum',
            'type': 'صحف (مقالات رأي)',
            'names': f'كتّاب رأي في:\n{chr(10).join(NEWSPAPERS_LOCAL)}',
            'paid': True,
            'cost': 6000,  # من earned_media_list earned7 (5 × 1200)
            'notes': '5 مقالات رأي وتحليلية'
        },
        {
            'phase': 'momentum',
            'type': 'بودكاست',
            'names': '\n'.join(PODCASTS),
            'paid': True,
            'cost': 2100,  # من earned_media_list earned8 (3 × 700)
            'notes': '3 حلقات بودكاست'
        },
        {
            'phase': 'momentum',
            'type': 'مؤثرين Macro (200K+)',
            'names': '\n'.join(INFLUENCERS_MOMENTUM[:2]),
            'paid': True,
            'cost': 6000,  # من earned_media_list earned9 (2 × 3000)
            'notes': '2 مؤثر كبير - تأثير عالي'
        },
        {
            'phase': 'momentum',
            'type': 'إعلانات رقمية (ترسيخ)',
            'names': 'LinkedIn Ads: 2,000\nRetargeting Ads: 1,200',
            'paid': True,
            'cost': 3200,  # من paid9+paid11
            'notes': 'إعادة استهداف وترسيخ'
        },
        {
            'phase': 'momentum',
            'type': 'إعلام دولي (ترسيخ)',
            'names': 'المخبر الاقتصادي (YouTube)\n5.5 مليون مشترك',
            'paid': True,
            'cost': 1400,  # من paid_media_list paid10
            'notes': 'حلقة تحليلية عن المشروع'
        },
    ]
    
    row = 2
    for i, tool in enumerate(detailed_tools, 1):
        phase_id = tool['phase']
        phase_name = PHASE_NAMES.get(phase_id, phase_id)
        phase_color = PHASE_COLORS.get(phase_id, 'FFFFFFFF')
        
        data = [
            i,
            phase_name,
            tool['type'],
            tool['names'],
            'مدفوع 💰' if tool['paid'] else 'مجاني ✓',
            f"{tool['cost']:,}" if tool['cost'] > 0 else '-',
            tool['notes']
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
            
            if col == 2:
                cell.fill = PatternFill(start_color=phase_color, end_color=phase_color, fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            
            if col == 5:
                if tool['paid']:
                    cell.fill = PatternFill(start_color='FFFEF3C7', end_color='FFFEF3C7', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFD1FAE5', end_color='FFD1FAE5', fill_type='solid')
        
        row += 1
    
    # ملخص
    row += 2
    ws.cell(row=row, column=1, value="📊 ملخص الأدوات الإعلامية")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 1
    
    total_paid = sum(t['cost'] for t in detailed_tools if t['paid'])
    ws.cell(row=row, column=1, value="إجمالي تكاليف الأدوات المدفوعة")
    ws.cell(row=row, column=2, value=f"{total_paid:,} {currency}")
    ws.cell(row=row, column=2).font = Font(bold=True, color='DC1F27', size=12)
    
    column_widths = [5, 28, 25, 55, 15, 15, 35]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # ضبط ارتفاع الصفوف للنص المتعدد الأسطر
    for r in range(2, row):
        ws.row_dimensions[r].height = 80

def create_paid_media_sheet(ws, campaign):
    """إنشاء شيت الوسائل المدفوعة"""
    ws.sheet_view.rightToLeft = True
    
    quotation = campaign.get('quotation', {})
    currency = quotation.get('currency', 'USD')
    
    headers = ['#', 'المرحلة', 'الوسيلة', 'التفاصيل', f'التكلفة ({currency})']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell, 'FFDC1F27')
    
    paid_list = quotation.get('paid_media_list', [])
    
    row = 2
    for i, item in enumerate(paid_list, 1):
        phase_id = item.get('phase_id', '')
        phase_name = PHASE_NAMES.get(phase_id, phase_id)
        
        data = [
            i,
            phase_name,
            item.get('name', ''),
            get_paid_media_details(item),
            f"{item.get('cost', 0):,}"
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        
        row += 1
    
    # إجمالي
    row += 1
    ws.cell(row=row, column=1, value="الإجمالي")
    ws.cell(row=row, column=1).font = Font(bold=True)
    total = sum(item.get('cost', 0) for item in paid_list)
    ws.cell(row=row, column=5, value=f"{total:,} {currency}")
    ws.cell(row=row, column=5).font = Font(bold=True, color='DC1F27', size=12)
    
    column_widths = [5, 28, 40, 50, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def get_paid_media_details(item):
    """تفاصيل الوسائل المدفوعة"""
    name = item.get('name', '').lower()
    
    if 'sms' in name or 'email' in name:
        return 'Twilio SMS + Mailchimp + قاعدة بيانات مستثمرين'
    elif 'linkedin' in name:
        return 'LinkedIn Campaign Manager - استهداف صناعي ولوجستي'
    elif 'google' in name:
        return 'Google Display Network - إعلانات بانر وفيديو'
    elif 'meta' in name or 'facebook' in name:
        return 'Meta Ads Manager - Facebook + Instagram'
    elif 'newswire' in name:
        return f'{", ".join(INTERNATIONAL_MEDIA)} + 50+ صحيفة أخرى'
    elif 'دولية' in name:
        return '\n'.join(INTERNATIONAL_MEDIA)
    else:
        return '-'

def create_earned_media_sheet(ws, campaign):
    """إنشاء شيت الإعلام الموازي"""
    ws.sheet_view.rightToLeft = True
    
    quotation = campaign.get('quotation', {})
    currency = quotation.get('currency', 'USD')
    
    headers = ['#', 'المرحلة', 'نوع الإعلام', 'الوسائل التفصيلية', 'الكمية', f'سعر الوحدة ({currency})', f'الإجمالي ({currency})']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell, 'FF10B981')
    
    earned_list = quotation.get('earned_media_list', [])
    
    row = 2
    grand_total = 0
    for i, item in enumerate(earned_list, 1):
        phase_id = item.get('phase_id', '')
        phase_name = PHASE_NAMES.get(phase_id, phase_id)
        quantity = item.get('quantity', 1)
        unit_price = item.get('unit_price', 0)
        total = quantity * unit_price
        grand_total += total
        
        data = [
            i,
            phase_name,
            item.get('type', ''),
            get_earned_media_names(item),
            quantity,
            f"{unit_price:,}",
            f"{total:,}"
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        
        row += 1
    
    # إجمالي
    row += 1
    ws.cell(row=row, column=1, value="الإجمالي")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=7, value=f"{grand_total:,} {currency}")
    ws.cell(row=row, column=7).font = Font(bold=True, color='10B981', size=12)
    
    column_widths = [5, 28, 20, 55, 10, 18, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    for r in range(2, row):
        ws.row_dimensions[r].height = 60

def get_earned_media_names(item):
    """أسماء وسائل الإعلام الموازي"""
    item_type = item.get('type', '').lower()
    name = item.get('name', '').lower()
    
    if 'صحافة' in item_type or 'صحف' in name:
        if 'رأي' in name:
            return '\n'.join(NEWSPAPERS_LOCAL)
        return '\n'.join(NEWSPAPERS_LOCAL[:3])
    elif 'مواقع' in item_type:
        return '\n'.join(NEWS_WEBSITES[:5])
    elif 'تلفزيون' in item_type:
        return '\n'.join(TV_CHANNELS[:3])
    elif 'راديو' in item_type:
        return '\n'.join(RADIO_STATIONS[:3])
    elif 'مؤثرين' in item_type:
        if 'micro' in name:
            return 'مؤثرين صغار (10K-50K متابع)'
        elif 'mid' in name:
            return 'مؤثرين متوسطين (50K-200K متابع)'
        elif 'macro' in name:
            return '\n'.join(INFLUENCERS_MOMENTUM[:2])
    elif 'بودكاست' in item_type:
        return '\n'.join(PODCASTS)
    
    return '-'

def create_budget_sheet(ws, campaign):
    """إنشاء شيت الميزانية"""
    ws.sheet_view.rightToLeft = True
    
    budget = campaign.get('budget', {})
    quotation = campaign.get('quotation', {})
    
    ws.merge_cells('A1:D1')
    ws['A1'] = "💰 توزيع الميزانية التفصيلي"
    ws['A1'].font = Font(bold=True, size=16, color='1E3A5F')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A3'] = "الميزانية الإجمالية:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = f"{quotation.get('grand_total', budget.get('total', 0)):,} {quotation.get('currency', budget.get('currency', 'USD'))}"
    ws['B3'].font = Font(bold=True, size=14, color='DC1F27')
    
    headers = ['الفئة', 'المبلغ', 'النسبة']
    row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    
    row = 6
    breakdown = budget.get('breakdown', [])
    for item in breakdown:
        ws.cell(row=row, column=1, value=item.get('category', ''))
        apply_cell_style(ws.cell(row=row, column=1))
        
        ws.cell(row=row, column=2, value=f"{item.get('amount', 0):,} USD")
        apply_cell_style(ws.cell(row=row, column=2))
        
        ws.cell(row=row, column=3, value=f"{item.get('percentage', 0)}%")
        apply_cell_style(ws.cell(row=row, column=3))
        
        row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

def main():
    """الدالة الرئيسية"""
    print("=" * 60)
    print("   📊 تصدير خطة الحملة إلى Excel - النسخة التفصيلية")
    print("=" * 60)
    
    print("\n⏳ جاري تحميل بيانات الحملة...")
    campaign = load_campaign_data()
    
    if not campaign:
        print("❌ لم يتم العثور على حملة ALIC")
        return
    
    print(f"✓ تم تحميل حملة: {campaign['basic_info']['name']}")
    print(f"  📦 عدد المنتجات: {len(campaign.get('communication_products', []))}")
    print(f"  🛠️  عدد الأدوات: {len(campaign.get('media_tools', []))}")
    
    print("\n⏳ جاري إنشاء ملف Excel التفصيلي...")
    output_path = create_excel_report(campaign)
    
    print("\n📋 الشيتات المُنشأة:")
    print("   1️⃣  نظرة عامة")
    print("   2️⃣  خطة المنتجات التفصيلية")
    print("   3️⃣  الأدوات الإعلامية التفصيلية (بالأسماء)")
    print("   4️⃣  الوسائل المدفوعة")
    print("   5️⃣  الإعلام الموازي")
    print("   6️⃣  الميزانية")
    
    print("\n" + "=" * 60)
    print("   ✅ تم بنجاح!")
    print("=" * 60)

if __name__ == '__main__':
    main()
