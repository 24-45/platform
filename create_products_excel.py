#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
سكريبت إنشاء ملف Excel للمنتجات والأدوات الإعلامية للحملة
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# إنشاء ملف Excel جديد
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "المنتجات الإعلامية"

# تعيين اتجاه الورقة من اليمين لليسار
ws.sheet_view.rightToLeft = True

# تعريف الأنماط
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

phase1_fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
phase2_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
phase3_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")

cell_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# العناوين
headers = [
    "المرحلة",
    "رقم المنتج",
    "اسم المنتج",
    "فكرة المنتج",
    "تكلفة الترويج المدفوع",
    "الإعلام الموازي",
    "تاريخ النشر المتوقع"
]

# كتابة العناوين
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# بيانات المنتجات - مستخرجة بدقة من المنتجات والأدوات الإعلامية للحملة
products = [
    # المرحلة الأولى: التأسيس والإطلاق (5 منتجات)
    {
        "phase": "المرحلة الأولى: التأسيس والإطلاق",
        "num": 1,
        "name": "إعلان تأسيس شركة ALIC",
        "idea": "إعلان رسمي عن تأسيس أول شركة أردنية متخصصة في الاستثمار العقاري بالمملكة العربية السعودية، مع التركيز على الشراكة الاستراتيجية مع Nobles Properties",
        "paid_cost": "لا يوجد",
        "parallel_media": "• صحف: الغد، المملكة، عمون، رؤيا، الدستور، سرايا، Jordan Times، جفرا، الرأي\n• قنوات: التلفزيون الأردني، رؤيا، المملكة\n• مؤثرون: مازن الساكت، فتحي الجغبير، نائل الكباريتي",
        "publish_date": "الأسبوع الأول من المرحلة الأولى"
    },
    {
        "phase": "المرحلة الأولى: التأسيس والإطلاق",
        "num": 2,
        "name": "جلسة حوارية مع عمر عايش",
        "idea": "جلسة حوارية معمقة مع المدير التنفيذي عمر عايش حول رؤية الشركة والفرص الاستثمارية في السوق السعودي",
        "paid_cost": "لا يوجد",
        "parallel_media": "• برنامج مال وأعمال (التلفزيون الأردني)\n• قناة المملكة\n• قناة رؤيا\n• صحيفة الغد",
        "publish_date": "الأسبوع الثاني من المرحلة الأولى"
    },
    {
        "phase": "المرحلة الأولى: التأسيس والإطلاق",
        "num": 3,
        "name": "فيديو تعريفي بالسوق السعودي",
        "idea": "فيديو احترافي يشرح مزايا الاستثمار العقاري في السوق السعودي وفرص رؤية 2030",
        "paid_cost": "5,000 - 10,000 دينار (حملة ترويجية على السوشيال ميديا)",
        "parallel_media": "• قنوات التواصل: Nobles Properties، الشهد، ALIC\n• منصات دولية: المخبر الاقتصادي، CNBC عربية",
        "publish_date": "الأسبوع الثالث من المرحلة الأولى"
    },
    {
        "phase": "المرحلة الأولى: التأسيس والإطلاق",
        "num": 4,
        "name": "سلسلة مقالات اقتصادية",
        "idea": "سلسلة من 5 مقالات متخصصة تتناول جوانب مختلفة من الاستثمار العقاري في السعودية",
        "paid_cost": "لا يوجد",
        "parallel_media": "• صحف: الغد، الدستور، الرأي، Jordan Times\n• كتّاب رأي: سلامة الدرعاوي، إبراهيم سيف، حسام عايش",
        "publish_date": "على مدار المرحلة الأولى (مقال أسبوعياً)"
    },
    {
        "phase": "المرحلة الأولى: التأسيس والإطلاق",
        "num": 5,
        "name": "إنفوجرافيك الفرص الاستثمارية",
        "idea": "سلسلة إنفوجرافيك توضح الأرقام والإحصائيات المتعلقة بالسوق العقاري السعودي",
        "paid_cost": "2,000 - 3,000 دينار (ترويج مدفوع)",
        "parallel_media": "• قنوات التواصل الاجتماعي\n• Property Finder\n• مواقع إخبارية",
        "publish_date": "الأسبوع الرابع من المرحلة الأولى"
    },
    
    # المرحلة الثانية: البناء والتوسع (8 منتجات)
    {
        "phase": "المرحلة الثانية: البناء والتوسع",
        "num": 6,
        "name": "قصص نجاح المستثمرين",
        "idea": "سلسلة فيديوهات قصيرة تعرض تجارب حقيقية لمستثمرين أردنيين نجحوا في السوق السعودي",
        "paid_cost": "3,000 - 5,000 دينار (لكل قصة)",
        "parallel_media": "• قنوات التواصل: Nobles Properties، ALIC\n• مؤثرون: موسى الساكت، حسن عبد الله\n• صحف: الغد، عمون",
        "publish_date": "الأسابيع 1-4 من المرحلة الثانية"
    },
    {
        "phase": "المرحلة الثانية: البناء والتوسع",
        "num": 7,
        "name": "ندوة استثمارية متخصصة",
        "idea": "ندوة حضورية وافتراضية تجمع خبراء ومستثمرين لمناقشة فرص السوق السعودي",
        "paid_cost": "15,000 - 20,000 دينار (تنظيم وترويج)",
        "parallel_media": "• تغطية: التلفزيون الأردني، رؤيا، المملكة\n• صحف: جميع الصحف الرئيسية\n• متحدثون: عمر عايش، أحمد مرعي، رشيد جعارة\n• ضيوف: مازن الساكت، فتحي الجغبير، نائل الكباريتي",
        "publish_date": "الأسبوع الثالث من المرحلة الثانية"
    },
    {
        "phase": "المرحلة الثانية: البناء والتوسع",
        "num": 8,
        "name": "فيلم سينمائي قصير",
        "idea": "فيلم وثائقي قصير (5-7 دقائق) يعرض رحلة مستثمر أردني في السوق السعودي",
        "paid_cost": "25,000 - 35,000 دينار (إنتاج وترويج)",
        "parallel_media": "• عرض: قنوات التلفزيون، يوتيوب\n• منصات دولية: CNBC عربية، Bloomberg الشرق، Sky News عربية\n• مهرجانات إعلامية",
        "publish_date": "الأسبوع الخامس من المرحلة الثانية"
    },
    {
        "phase": "المرحلة الثانية: البناء والتوسع",
        "num": 9,
        "name": "بودكاست استثماري",
        "idea": "سلسلة حلقات بودكاست تستضيف خبراء ومستثمرين للحديث عن تجاربهم",
        "paid_cost": "2,000 - 4,000 دينار (لكل حلقة)",
        "parallel_media": "• بودكاست بذرة (أنس شبارو)\n• المخبر الاقتصادي\n• منصات البودكاست: Spotify، Apple Podcasts\n• ضيوف: السفراء والمؤثرون",
        "publish_date": "حلقة كل أسبوعين خلال المرحلة الثانية"
    },
    {
        "phase": "المرحلة الثانية: البناء والتوسع",
        "num": 10,
        "name": "حملة العلاقات العامة الرقمية",
        "idea": "حملة متكاملة على منصات التواصل الاجتماعي لتعزيز الوعي بالعلامة التجارية",
        "paid_cost": "20,000 - 30,000 دينار (شهرياً)",
        "parallel_media": "• فيسبوك، إنستغرام، لينكد إن، تويتر\n• مؤثرون: باسم الزعبي، مازن العمري، فارس الصايغ\n• Property Finder",
        "publish_date": "مستمرة طوال المرحلة الثانية"
    },
    {
        "phase": "المرحلة الثانية: البناء والتوسع",
        "num": 11,
        "name": "تقرير السوق الربعي",
        "idea": "تقرير تحليلي شامل عن أداء السوق العقاري السعودي والفرص المتاحة",
        "paid_cost": "لا يوجد",
        "parallel_media": "• صحف: الغد، الدستور، Jordan Times\n• Forbes Middle East\n• مواقع اقتصادية متخصصة",
        "publish_date": "نهاية المرحلة الثانية"
    },
    {
        "phase": "المرحلة الثانية: البناء والتوسع",
        "num": 12,
        "name": "جولة إعلامية سعودية",
        "idea": "جولة للمتحدثين الرسميين في وسائل الإعلام السعودية لتعزيز الحضور",
        "paid_cost": "30,000 - 40,000 دينار (سفر وتنظيم)",
        "parallel_media": "• قنوات سعودية: العربية، الإخبارية، MBC\n• صحف: الرياض، عكاظ، الاقتصادية\n• متحدثون: عمر عايش، أحمد مرعي",
        "publish_date": "الأسبوع السادس من المرحلة الثانية"
    },
    {
        "phase": "المرحلة الثانية: البناء والتوسع",
        "num": 13,
        "name": "ورش عمل تثقيفية",
        "idea": "ورش عمل مجانية لتثقيف المستثمرين المحتملين حول آليات الاستثمار",
        "paid_cost": "5,000 - 8,000 دينار (لكل ورشة)",
        "parallel_media": "• تغطية صحفية: الغد، عمون، رؤيا\n• بث مباشر على السوشيال ميديا\n• مقدمون: رشيد جعارة، أحمد مرعي",
        "publish_date": "ورشة كل أسبوعين في المرحلة الثانية"
    },
    
    # المرحلة الثالثة: التعزيز والاستدامة (7 منتجات)
    {
        "phase": "المرحلة الثالثة: التعزيز والاستدامة",
        "num": 14,
        "name": "تقرير الإنجازات السنوي",
        "idea": "تقرير شامل يوثق إنجازات الشركة خلال العام الأول",
        "paid_cost": "لا يوجد",
        "parallel_media": "• صحف: جميع الصحف الرئيسية\n• قنوات: التلفزيون الأردني، رؤيا، المملكة\n• Forbes Middle East",
        "publish_date": "الأسبوع الأول من المرحلة الثالثة"
    },
    {
        "phase": "المرحلة الثالثة: التعزيز والاستدامة",
        "num": 15,
        "name": "حفل تكريم المستثمرين",
        "idea": "حفل سنوي لتكريم المستثمرين المتميزين وشركاء النجاح",
        "paid_cost": "50,000 - 70,000 دينار",
        "parallel_media": "• تغطية شاملة: جميع القنوات والصحف\n• مؤثرون: جميع السفراء والمؤثرين\n• بث مباشر",
        "publish_date": "الأسبوع الثاني من المرحلة الثالثة"
    },
    {
        "phase": "المرحلة الثالثة: التعزيز والاستدامة",
        "num": 16,
        "name": "سلسلة لقاءات تلفزيونية",
        "idea": "سلسلة لقاءات مع المتحدثين الرسميين في البرامج الصباحية والاقتصادية",
        "paid_cost": "لا يوجد",
        "parallel_media": "• التلفزيون الأردني (برنامج مال وأعمال)\n• رؤيا (يسعد صباحك)\n• المملكة\n• متحدثون: عمر عايش، أحمد مرعي، رشيد جعارة",
        "publish_date": "على مدار المرحلة الثالثة"
    },
    {
        "phase": "المرحلة الثالثة: التعزيز والاستدامة",
        "num": 17,
        "name": "جولة إذاعية Radio Tour",
        "idea": "جولة على الإذاعات الأردنية الرئيسية للترويج والتوعية",
        "paid_cost": "لا يوجد",
        "parallel_media": "• هلا FM\n• Play 99.6\n• روتانا FM\n• Energy FM\n• Mazaj FM\n• متحدثون: عمر عايش، أحمد مرعي",
        "publish_date": "الأسبوع الثالث من المرحلة الثالثة"
    },
    {
        "phase": "المرحلة الثالثة: التعزيز والاستدامة",
        "num": 18,
        "name": "حملة شهادات العملاء",
        "idea": "حملة فيديوهات قصيرة تعرض شهادات العملاء الراضين",
        "paid_cost": "10,000 - 15,000 دينار",
        "parallel_media": "• السوشيال ميديا\n• موقع الشركة\n• Property Finder\n• مؤثرون للمشاركة",
        "publish_date": "الأسابيع 3-6 من المرحلة الثالثة"
    },
    {
        "phase": "المرحلة الثالثة: التعزيز والاستدامة",
        "num": 19,
        "name": "دليل المستثمر الشامل",
        "idea": "كتيب إلكتروني شامل يغطي كل ما يحتاجه المستثمر للبدء",
        "paid_cost": "3,000 - 5,000 دينار (تصميم وترويج)",
        "parallel_media": "• توزيع: البريد الإلكتروني، SMS\n• موقع الشركة\n• Property Finder\n• السفارة السعودية",
        "publish_date": "الأسبوع الرابع من المرحلة الثالثة"
    },
    {
        "phase": "المرحلة الثالثة: التعزيز والاستدامة",
        "num": 20,
        "name": "خطة المرحلة القادمة",
        "idea": "إعلان استراتيجية الشركة للعام القادم والمشاريع الجديدة",
        "paid_cost": "لا يوجد",
        "parallel_media": "• مؤتمر صحفي\n• جميع وسائل الإعلام\n• Forbes Middle East\n• متحدث: عمر عايش",
        "publish_date": "الأسبوع الأخير من المرحلة الثالثة"
    }
]

# كتابة البيانات
row = 2
for product in products:
    # تحديد لون الصف حسب المرحلة
    if "الأولى" in product["phase"]:
        fill = phase1_fill
    elif "الثانية" in product["phase"]:
        fill = phase2_fill
    else:
        fill = phase3_fill
    
    ws.cell(row=row, column=1, value=product["phase"]).fill = fill
    ws.cell(row=row, column=2, value=product["num"]).fill = fill
    ws.cell(row=row, column=3, value=product["name"]).fill = fill
    ws.cell(row=row, column=4, value=product["idea"]).fill = fill
    ws.cell(row=row, column=5, value=product["paid_cost"]).fill = fill
    ws.cell(row=row, column=6, value=product["parallel_media"]).fill = fill
    ws.cell(row=row, column=7, value=product["publish_date"]).fill = fill
    
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.alignment = cell_alignment
        cell.border = border
    
    row += 1

# تعديل عرض الأعمدة
column_widths = [35, 12, 30, 60, 25, 60, 35]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# تعديل ارتفاع الصفوف
ws.row_dimensions[1].height = 30
for r in range(2, row):
    ws.row_dimensions[r].height = 80

# حفظ الملف
output_file = "ALIC_Products_Media_Plan.xlsx"
wb.save(output_file)
print(f"✅ تم إنشاء ملف Excel بنجاح: {output_file}")
print(f"📊 عدد المنتجات: {len(products)}")
print(f"   - المرحلة الأولى: 5 منتجات")
print(f"   - المرحلة الثانية: 8 منتجات")
print(f"   - المرحلة الثالثة: 7 منتجات")
